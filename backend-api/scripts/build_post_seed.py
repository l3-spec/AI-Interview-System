import json
import subprocess
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORD_DIR = ROOT.parent / "word"
OUTPUT_JSON = Path(__file__).resolve().parent / "post_seed.json"

posts = []


def extract_docx(doc_path: Path):
    try:
        raw = subprocess.check_output([
            "textutil",
            "-convert",
            "txt",
            "-stdout",
            str(doc_path)
        ])
        text = raw.decode("utf-8", errors="ignore")
    except subprocess.CalledProcessError as exc:
        print(f"[warn] failed to read {doc_path.name}: {exc}")
        return None

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return None

    title = lines[0]
    content_lines = lines[1:] or lines
    content = "\n".join(content_lines).strip()

    # 基于文件名拆分标签
    stem = doc_path.stem
    tag_parts = [p.strip() for p in stem.replace('：', ':').split(':') if p.strip()]
    tags = [p for p in tag_parts if len(p) > 0]

    return {
        "title": title,
        "content": content,
        "tags": tags,
        "source": doc_path.name,
        "status": "PUBLISHED"
    }


def extract_from_excel(excel_path: Path):
    wb = load_workbook(excel_path)
    sheet_posts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            # 第三列包含“标题：…内容：…”格式
            text = row[2] if len(row) >= 3 else None
            if not text or not isinstance(text, str):
                continue
            raw = text.strip().replace('\u3000', ' ')
            if not raw:
                continue

            title = raw
            content = ""
            if "内容：" in raw:
                title_part, content_part = raw.split("内容：", 1)
                title = title_part.replace("标题：", "").strip()
                content = content_part.strip()
            else:
                if "标题：" in raw:
                    title = raw.replace("标题：", "").strip()

            tags = []
            if len(row) >= 1 and row[0]:
                tags.append(str(row[0]).strip())
            if len(row) >= 2 and row[1]:
                tags.append(str(row[1]).strip())

            sheet_posts.append({
                "title": title,
                "content": content or title,
                "tags": [t for t in tags if t],
                "source": f"{excel_path.name}:{ws.title}",
                "status": "PUBLISHED"
            })
    return sheet_posts


def main():
    # Docx only (按需求仅保留 Word 全文帖子)
    for doc_path in sorted(WORD_DIR.glob("*.docx")):
        extracted = extract_docx(doc_path)
        if extracted:
            posts.append(extracted)

    # 去重标题
    title_seen = set()
    unique_posts = []
    for p in posts:
        t = p["title"]
        if t in title_seen:
            continue
        title_seen.add(t)
        unique_posts.append(p)

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(unique_posts, f, ensure_ascii=False, indent=2)

    print(f"Generated {len(unique_posts)} posts -> {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
